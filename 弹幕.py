import requests  # 引入网站数据获取库
import re  # 引入正则表达式库
import warnings  # 引入忽略警告库
import stylecloud  # 引入词云图生成库
import openpyxl  # 引入写入Excel文件库
import tqdm  # 引入进度条库

# 请求头
headers = {
    "cookie": 'buvid3=25FC892C-5144-4826-D1E6-E283CED9D84A93079infoc; b_nut=1694088893; i-wanna-go-back=-1; b_ut=7; _uuid=52AED4110-329E-6F34-A544-EAC98A1D549C92631infoc; buvid_fp=de94f45ec0f8162b0d5fabcf4db29033; buvid4=EB5E5E44-863B-741C-D9B5-36C4A4FC53EB95085-023090720-Sfw%2Bq8N2F3%2FX4sk3WHE9UA%3D%3D; DedeUserID=1680065450; DedeUserID__ckMd5=2678661500128859; header_theme_version=CLOSE; home_feed_column=5; CURRENT_FNVAL=4048; rpdid=|(u))kklm)Ru0J\'uYmR|muRuJ; SESSDATA=224ec5e5%2C1709991813%2C41543%2A92CjAHnfsvJC3Xk3jBFWUQEQkkPYFm-WVdbeg2_Ul7FofI0HdG9BQD9avWtuGYS24RjxgSVkdER1Zub1p2RXBXZHZqRXFVQTJ3NHVMR1NrYU9weGt2SExSejdTakZhb1hZMVBHeXdheVl2YUdwazI4eFhBcV9UV3BSbV9IUndHSGxQTGQ5N25JRnZ3IIEC; bili_jct=634f158dc0e4a5e37b624bbf43de430d; sid=72h10c7x; bili_ticket=eyJhbGciOiJIUzI1NiIsImtpZCI6InMwMyIsInR5cCI6IkpXVCJ9.eyJleHAiOjE2OTQ2OTkzNzksImlhdCI6MTY5NDQ0MDE3OSwicGx0IjotMX0.50k57T_Dp3XRHpbvk_9JXv2GoA4s-7A7sxUlRTxsG9s; bili_ticket_expires=1694699379; bp_video_offset_1680065450=840092897296515105; innersign=0; b_lsid=6478A12B_18A8C9BAB29; browser_resolution=1488-318; PVID=1',
    "origin": 'https://www.bilibili.com',
    "referer": 'https://www.bilibili.com/',
    "user-agent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36 Edg/116.0.1938.62'
}


# 网站属性
def get_params(page):
    try:
        params = {
            '__refresh__': 'true',
            '_extra': '',
            'context': '',
            'page': page,
            'page_size': 30,
            'from_source': '',
            'from_spmid': '333.337',
            'platform': 'pc',
            'highlight': '1',
            'single_column': '0',
            'keyword': '日本核污染水排海',
            'qv_id': '806wcU3HaRokbwxU01nB6AsfUHqOGjT3',
            'source_tag': '3',
            'gaia_vtoken': '',
            'category_id': '',
            'search_type': 'video',
            'dynamic_offset': '90',
            'web_location': '1430654',
            'w_rid': 'f8103d5bcc9d58e676e2d94485f64d9d',
            'wts': '1694595911'
        }
    # 异常处理
    except Exception as e:
        print("get_params出错：", e)
    return params


# 请求BV号
def get_video_bvids():
    try:
        video_bvids = []
        for page in tqdm.tqdm(range(1, 11)):
            # 获取每页的params，以便获取bv号
            params = get_params(page)
            url = 'https://api.bilibili.com/x/web-interface/wbi/search/type'
            response = requests.get(url=url, headers=headers, params=params)
            # 转换json类型，便于获取bvid号
            res = response.json()
            # 通过json类型查询并获取bv号
            bvids = res['data']['result']
            for bvid in tqdm.tqdm(bvids):
                video_bvids.append(bvid['bvid'])
    # 异常处理
    except Exception as e:
        print("get_video_bvids出错：", e)
    return video_bvids


# 根据bvid请求得到cid

# 请求cid号
def get_cid(bvid):
    try:
        # 视频地址：https://www.bilibili.com/video/BV1PK4y1b7dt?t=1
        url = f'https://api.bilibili.com/x/player/pagelist?bvid={bvid}'
        # 转换接送类型，便于获取cid号
        res = requests.get(url).json()
        # 将获取的网页json编码字符串转换为python对象，字典类型
        cid = res["data"][0]['cid']
    # 异常处理
    except Exception as e:
        print("get_cid出错：", e)
    return cid


# 生成词云
def get_wordcloud():
    try:
    # 删除重复元素
        my_long_list = ["保护海洋", "见证历史", "反对排放", "支持", "保护地球", "好", "坚决抵制日本核污水排海", "保护地球保护海洋", "世界可以没有日本", "但不能没有海洋", "发货",
                        "坚决抵制", "年"]
        stylecloud.gen_stylecloud(file_path='./弹幕.txt',  # 文件地址
                                  icon_name='fas fa-biohazard',  # 词云样式
                                  palette='colorbrewer.diverging.Spectral_11',  # 词云图颜色
                                  background_color='black',  # 背景颜色
                                  gradient='horizontal',  # 梯度的方向
                                  font_path="msyh.ttc",  # 字体
                                  size=2048,  # 词云图尺寸
                                  custom_stopwords=my_long_list,  # 自定义停用词列表
                                  output_name="./弹幕.png",  # 输出文件名
                                  )
    # 异常处理
    except Exception as e:
        print("get_wordcloud出错：", e)

# 将弹幕保存在Excel文件中
def write_danmu_excel():
    try:
        # 创建工作簿
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # 打开文件获取弹幕
        file = open('./弹幕.txt', 'r', encoding='utf-8')
        lines = file.readlines()

        # 将弹幕写入Excel文件中
        for line in lines:
            sheet.append([line.strip()])

        # 保存文件
        workbook.save('./弹幕.xlsx')
    # 异常处理
    except Exception as e:
        print("write_danmu_excel出错：", e)


# 输出频次前二十的弹幕并且生成弹幕统计的Excel
def sort_danmutop20():
    try:
        # 创建工作簿
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # 给Excel文件第一行两列赋值
        sheet.cell(row=1, column=1).value = '弹幕'
        sheet.cell(row=1, column=2).value = '频率'
        # 打开文件获取弹幕
        with open('./弹幕.txt', 'r', encoding='utf-8') as data:
            danmus = data.read().splitlines()
        danmu_dic = {}

        # 判断如果弹幕在字典里，则值加1，否则在字典中加入该弹幕，值赋为1
        for danmu in danmus:
            if danmu in danmu_dic:
                danmu_dic[danmu] += 1
            else:
                danmu_dic[danmu] = 1

        # 根据弹幕出现频次排序弹幕
        danmu_dic = sorted(danmu_dic.items(), key=lambda x: x[1], reverse=True)

        # 生成弹幕统计的Excel
        for i in range(len(danmu_dic)):
            sheet.append([danmu_dic[i][0], danmu_dic[i][1]])

        # 输出出现频率前二十的弹幕
        for i in range(20):
            print(str(danmu_dic[i][0]), ":", str(danmu_dic[i][1]))

        # 保存文件
        workbook.save('./弹幕统计.xlsx')
    # 异常处理
    except Exception as e:
        print("sort_danmutop20", e)


def get_danmu():
    try:
        # 获取bv号
        bvids = get_video_bvids()
        for bvid in tqdm.tqdm(bvids):

            # 弹幕地址
            url = f'http://comment.bilibili.com/{get_cid(bvid)}.xml'
            response = requests.get(url=url, headers=headers)

            # 转换编码类型便于查找
            response.encoding = 'utf-8'

            # 正则表达式，数据解析，提取弹幕内容
            data = re.findall('<d p=".*?">(.*?)</d>', response.text)

            for index in data:
                # 保存数据在txt文件中便于生成词云图
                # 写入txt文件
                with open("./弹幕.txt", 'a', encoding='utf-8') as file:
                    file.write(index)
                    file.write('\n')
    # 异常处理
    except Exception as e:
        print("get_danmu", e)


if __name__ == '__main__':
    # 忽略警告
    warnings.filterwarnings("ignore")

    # 获取弹幕
    get_danmu()

    # 输出弹幕频次前二十的Excel文件并把弹幕统计结果写入Excel中
    sort_danmutop20()
    # 将弹幕保存在Excel文件中
    write_danmu_excel()
    # 构建词云图
    get_wordcloud()
